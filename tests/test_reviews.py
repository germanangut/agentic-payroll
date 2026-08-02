import pandas as pd
import pytest
from openpyxl import load_workbook

from agentic_nomina.reporting.excel import write_report
from agentic_nomina.reporting.reviews import apply_review_ledger, load_review_ledger


def _exceptions(actual_value: float = 90.0) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "module": "SOCIAL_SECURITY",
                "period_label": "MONTH",
                "employee_id": "123",
                "employee_name": "Empleado Uno",
                "control": "health",
                "expected_value": 100.0,
                "actual_value": actual_value,
                "difference": 100.0 - actual_value,
                "severity": "REVIEW",
                "rule_id": "social_security_baseline",
                "rule_version": "1.0",
                "rule_status": "PROVISIONAL",
                "source_file": "pila.xlsx",
                "source_sheet": "PILA",
                "source_row": 2,
                "notes": "test",
            }
        ]
    )


def test_review_decision_is_preserved_for_unchanged_finding() -> None:
    pending = apply_review_ledger(_exceptions())
    reviews = pending[
        [
            "exception_id",
            "review_status",
            "review_decision",
            "reviewer",
            "reviewed_at",
            "review_notes",
        ]
    ].copy()
    reviews.loc[0, ["review_status", "review_decision", "reviewer", "reviewed_at"]] = [
        "RESUELTO",
        "CONFIRMADO",
        "analista.nomina",
        "2026-08-02",
    ]

    resolved = apply_review_ledger(_exceptions(), reviews)

    assert resolved.loc[0, "review_status"] == "RESUELTO"
    assert resolved.loc[0, "reviewer"] == "analista.nomina"


def test_changed_finding_rejects_stale_review() -> None:
    previous = apply_review_ledger(_exceptions())
    reviews = previous[
        [
            "exception_id",
            "review_status",
            "review_decision",
            "reviewer",
            "reviewed_at",
            "review_notes",
        ]
    ]

    with pytest.raises(ValueError, match="no longer current"):
        apply_review_ledger(_exceptions(actual_value=80.0), reviews)


def test_resolved_review_requires_auditable_fields() -> None:
    pending = apply_review_ledger(_exceptions())
    reviews = pending[
        [
            "exception_id",
            "review_status",
            "review_decision",
            "reviewer",
            "reviewed_at",
            "review_notes",
        ]
    ].copy()
    reviews.loc[0, "review_status"] = "RESUELTO"

    with pytest.raises(ValueError, match="require decision, reviewer and reviewed_at"):
        apply_review_ledger(_exceptions(), reviews)


def test_report_exports_reusable_review_sheet(tmp_path) -> None:
    output = tmp_path / "reconciliation.xlsx"
    employees = {
        "Q1": pd.DataFrame(
            [
                {
                    "employee_id": "123",
                    "employee_name_list": "Empleado Uno",
                    "employee_name_payroll": "Empleado Uno",
                    "employee_status": "E",
                    "in_employee_list": True,
                    "in_payroll": False,
                    "outcome": "MISSING_IN_PAYROLL",
                    "severity": "WARNING",
                }
            ]
        )
    }
    social = pd.DataFrame(
        [
            {
                "employee_id": "123",
                "employee_name_payroll": "Empleado Uno",
                "health_severity": "OK",
                "pension_severity": "OK",
                "days_severity": "OK",
            }
        ]
    )

    write_report(output, employees, social)

    workbook = load_workbook(output, read_only=True)
    assert "Revisiones" in workbook.sheetnames
    ledger = load_review_ledger(output)
    assert ledger.loc[0, "review_status"] == "PENDIENTE"
    assert ledger.loc[0, "exception_id"].startswith("EXC-")
