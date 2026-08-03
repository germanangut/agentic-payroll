import pandas as pd
import pytest
from openpyxl import load_workbook

from agentic_nomina.reporting.excel import write_report
from agentic_nomina.reporting.rules import (
    apply_rule_ledger,
    require_approved_financial_rules,
)


def _registry() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "rule_id": "financial_rule",
                "rule_version": "2.0",
                "rule_name": "Regla financiera",
                "active": True,
                "financial": True,
            },
            {
                "rule_id": "non_financial_rule",
                "rule_version": "1.0",
                "rule_name": "Regla operativa",
                "active": True,
                "financial": False,
            },
        ]
    )


def _approval(rule_id: str = "financial_rule", version: str = "2.0") -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "rule_id": rule_id,
                "rule_version": version,
                "estado_aprobacion": "APROBADA",
                "responsable_aprobacion": "lider.nomina",
                "fecha_aprobacion": "2026-08-02",
                "evidencia_aprobacion": "ACTA-42",
            }
        ]
    )


def test_rule_approval_is_version_specific_and_required_in_strict_mode() -> None:
    rules = apply_rule_ledger(_registry(), _approval())
    require_approved_financial_rules(rules)

    stale = _approval(version="1.0")
    with pytest.raises(ValueError, match="obsolete approvals"):
        apply_rule_ledger(_registry(), stale)


def test_rule_ledger_rejects_duplicate_and_unknown_approvals() -> None:
    duplicated = pd.concat([_approval(), _approval()], ignore_index=True)
    with pytest.raises(ValueError, match="duplicate approvals"):
        apply_rule_ledger(_registry(), duplicated)

    with pytest.raises(ValueError, match="unknown rules"):
        apply_rule_ledger(_registry(), _approval(rule_id="unknown_rule"))


def test_strict_mode_rejects_pending_or_incomplete_financial_approval() -> None:
    pending = apply_rule_ledger(_registry())
    with pytest.raises(ValueError, match="financial_rule@2.0"):
        require_approved_financial_rules(pending)

    incomplete = _approval()
    incomplete.loc[0, "evidencia_aprobacion"] = ""
    with pytest.raises(ValueError, match="require responsible person"):
        apply_rule_ledger(_registry(), incomplete)


def test_report_exports_rule_register(tmp_path) -> None:
    output = tmp_path / "reconciliation.xlsx"
    employees = {"Q1": pd.DataFrame([{"employee_id": "1", "severity": "OK"}])}
    social = pd.DataFrame(
        [
            {
                "employee_id": "1",
                "health_severity": "OK",
                "pension_severity": "OK",
                "days_severity": "OK",
            }
        ]
    )

    write_report(output, employees, social, rules=apply_rule_ledger(_registry(), _approval()))

    workbook = load_workbook(output, read_only=True)
    assert "Reglas" in workbook.sheetnames
    values = list(workbook["Reglas"].values)
    assert "estado_aprobacion" in values[0]
    assert "APROBADA" in values[1]
