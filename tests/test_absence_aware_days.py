import pandas as pd
from openpyxl import load_workbook

from agentic_nomina.reconciliation.absence_aware_days import explain_contributed_day_differences
from agentic_nomina.reporting.excel import write_report


def _social(days_difference: float) -> pd.DataFrame:
    return pd.DataFrame([{"employee_id": "100", "payroll_salary_days": 28.0, "health_days": 30.0, "days_difference": days_difference, "health_severity": "OK", "pension_severity": "OK", "days_severity": "REVIEW"}])


def _evidence(units: float = 2.0, absence_type: str = "INCAPACIDAD") -> pd.DataFrame:
    return pd.DataFrame([{"employee_id": "100", "period_label": "MONTH", "absence_type": absence_type, "units": units, "support_reference": "SYN-1", "evidence_status": "VALID", "unit": "DAYS"}])


def _config() -> dict[str, object]:
    return {"rule_id": "absence_aware_contributed_days", "rule_version": "1.0", "absence_types": ["INCAPACIDAD"], "comparable_units": ["DAYS"]}


def test_exact_absence_match_is_explanatory_and_preserves_sources() -> None:
    result = explain_contributed_day_differences(_social(-2), _evidence(), _config()).iloc[0]
    assert result["days_explanation_status"] == "EXPLICACION_PROVISIONAL"
    assert result["payroll_salary_days"] == 28
    assert result["health_days"] == 30
    assert not result["absence_financial_effect"]


def test_hours_or_partial_absence_requires_review() -> None:
    hours = _evidence()
    hours.loc[0, "unit"] = "HOURS"
    assert explain_contributed_day_differences(_social(-2), hours, _config()).iloc[0]["days_explanation_status"] == "REVIEW"
    assert explain_contributed_day_differences(_social(-2), _evidence(1), _config()).iloc[0]["unexplained_days_difference"] == -2


def test_synthetic_workbook_reopens_with_absence_traceability(tmp_path) -> None:
    social = explain_contributed_day_differences(_social(-2), _evidence(), _config())
    output = tmp_path / "synthetic.xlsx"
    employees = {"Q1": pd.DataFrame([{"employee_id": "100", "severity": "OK"}])}
    write_report(output, employees, social, absences=pd.DataFrame())
    workbook = load_workbook(output, read_only=True)
    assert {"Seguridad_Social", "Ausencias", "Excepciones", "Revisiones", "Casos_Empleado", "Reglas"} <= set(workbook.sheetnames)
    headers = [cell.value for cell in next(workbook["Seguridad_Social"].iter_rows(max_row=1))]
    assert "contributed_days_control_id" in headers
    assert "absence_financial_effect" in headers
