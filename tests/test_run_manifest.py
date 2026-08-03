import pandas as pd
import pytest
from openpyxl import load_workbook

from agentic_nomina.config import load_config
from agentic_nomina.reporting.excel import write_report
from agentic_nomina.run_manifest import (
    build_run_metadata,
    execution_frame,
    load_external_manifest,
    preflight_sources,
    resolve_run_contract,
    validate_period,
)
from agentic_nomina.service import run_baseline


def test_manifest_hashes_synthetic_source_without_path(tmp_path) -> None:
    source = tmp_path / "synthetic.txt"
    source.write_text("synthetic", encoding="utf-8")
    metadata, rows, _ = build_run_metadata(
        "Empresa Sintética", "2026-02", "RUN-ONE", {"pila": source}
    )
    assert metadata.run_id == "RUN-ONE"
    assert rows[0]["file_name"] == "pila.txt"
    assert "path" not in rows[0]
    assert len(rows[0]["sha256"]) == 64


def test_period_is_conservative() -> None:
    assert validate_period(None) == "NO_ESPECIFICADO"
    with pytest.raises(ValueError, match="YYYY-MM"):
        validate_period("2026-13")


def _config() -> dict[str, object]:
    return {"rule_governance": {"registry": [{"rule_id": "synthetic", "rule_version": "1.0", "active": True}]}}


def _required_sources(tmp_path) -> dict[str, object]:
    return {name: tmp_path / f"{name}.xlsx" for name in ("payroll_q1", "payroll_q2", "employees_q1", "employees_q2", "pila")}


def test_preflight_validates_pairs_extensions_and_duplicates(tmp_path) -> None:
    sources = _required_sources(tmp_path)
    for source in sources.values():
        source.write_bytes(b"synthetic")
    preflight_sources(sources, _config())
    sources["overtime_q1"] = tmp_path / "q1.xlsx"
    sources["overtime_q1"].write_bytes(b"synthetic")
    with pytest.raises(ValueError, match="declararse juntas"):
        preflight_sources(sources, _config())
    sources["overtime_q2"] = sources["overtime_q1"]
    with pytest.raises(ValueError, match="misma ruta"):
        preflight_sources(sources, _config())


def test_external_manifest_precedence_and_execution_are_path_free(tmp_path) -> None:
    manifest = tmp_path / "run.yml"
    manifest.write_text(
        "schema_version: '1.0'\nbusiness_period: '2026-01'\nrun_id: MANIFEST\nsources: {}\n",
        encoding="utf-8",
    )
    sources, period, run_id, diagnostics = resolve_run_contract(
        {}, period="2026-02", run_id="CLI", manifest_path=manifest
    )
    assert sources == {}
    assert (period, run_id) == ("2026-02", "CLI")
    assert len(diagnostics) == 2
    metadata, source_rows, preflight = build_run_metadata(
        "Empresa Sintética", period, run_id, {}, diagnostics=diagnostics
    )
    execution = execution_frame(
        metadata,
        source_rows,
        preflight,
        pd.DataFrame([{"rule_id": "synthetic", "rule_version": "1.0", "active": True, "financial": False, "estado_aprobacion": "PENDIENTE"}]),
    )
    assert set(execution["record_type"]) == {"CORRIDA", "DIAGNOSTICO", "REGLA"}
    assert str(tmp_path) not in execution.to_csv(index=False)


def test_invalid_manifest_and_active_rule_without_version_fail(tmp_path) -> None:
    malformed = tmp_path / "invalid.yml"
    malformed.write_text("schema_version: '2.0'\nsources: {}\n", encoding="utf-8")
    with pytest.raises(ValueError, match="schema_version"):
        load_external_manifest(malformed)
    source = tmp_path / "one.xlsx"
    source.write_bytes(b"synthetic")
    with pytest.raises(ValueError, match="rule_id y rule_version"):
        preflight_sources(
            {name: source for name in _required_sources(tmp_path)},
            {"rule_governance": {"registry": [{"rule_id": "bad", "active": True}]}},
        )


def test_execution_sheet_reopens_without_local_paths(tmp_path) -> None:
    output = tmp_path / "synthetic.xlsx"
    metadata, sources, diagnostics = build_run_metadata("Empresa Sintética", "2026-02", "RUN-ONE", {})
    execution = execution_frame(metadata, sources, diagnostics, pd.DataFrame())
    employees = {"Q1": pd.DataFrame([{"employee_id": "SYN-001", "severity": "OK"}])}
    social = pd.DataFrame([{"employee_id": "SYN-001", "health_severity": "OK", "pension_severity": "OK", "days_severity": "OK"}])
    write_report(output, employees, social, execution=execution)
    workbook = load_workbook(output, read_only=True)
    assert {"Ejecucion", "Excepciones", "Revisiones", "Casos_Empleado", "Reglas"} <= set(workbook.sheetnames)
    headers = [cell.value for cell in next(workbook["Ejecucion"].iter_rows(max_row=1))]
    assert {"run_id", "period", "execution_timestamp", "schema_version", "preflight_status"} <= set(headers)
    assert str(tmp_path) not in "".join(str(cell.value) for row in workbook["Ejecucion"].iter_rows() for cell in row)


def test_public_service_writes_auditable_execution_sheet(tmp_path) -> None:
    config = load_config("config/baseline.yml")
    payroll_headers = ["EMPLEADO", "NOMBRE EMPLEADO", "VAR", "SUELDO BASICO"]
    for label in ("q1", "q2"):
        pd.DataFrame([payroll_headers, ["9001", "Empleado Sintético", 30, 1_000_000]]).to_excel(
            tmp_path / f"payroll-{label}.xlsx", header=False, index=False, sheet_name="Hoja1"
        )
        pd.DataFrame([["NIT", "NOMBRE", "EST", "F. RETIRO"], ["9001", "Empleado Sintético", "ACTIVO", ""]]).to_excel(
            tmp_path / f"employees-{label}.xlsx", header=False, index=False, sheet_name="Hoja1"
        )
    pila_row = [None] * 70
    pila_row[4], pila_row[8], pila_row[35], pila_row[36], pila_row[38] = "9001", "Empleado Sintético", 30, 1_000_000, 80_000
    pila_row[42], pila_row[43], pila_row[46], pila_row[48], pila_row[50], pila_row[51] = 30, 1_000_000, 40_000, 30, 1_000_000, 0
    pila_row[55], pila_row[58], pila_row[61], pila_row[62], pila_row[69] = 30, 1_000_000, 0, 0, 120_000
    pd.DataFrame([pila_row]).to_excel(tmp_path / "pila.xlsx", header=False, index=False, sheet_name="Sheet1")
    output = tmp_path / "report.xlsx"
    results = run_baseline(
        payroll_q1_path=tmp_path / "payroll-q1.xlsx", payroll_q2_path=tmp_path / "payroll-q2.xlsx",
        employees_q1_path=tmp_path / "employees-q1.xlsx", employees_q2_path=tmp_path / "employees-q2.xlsx",
        pila_path=tmp_path / "pila.xlsx", output_path=output, config=config,
        business_period="2026-02", run_id="SYNTHETIC-RUN",
    )
    workbook = load_workbook(output, read_only=True)
    execution = pd.read_excel(output, sheet_name="Ejecucion")
    assert results["run_metadata"].iloc[0]["run_id"] == "SYNTHETIC-RUN"
    assert {"Ejecucion", "Seguridad_Social", "Excepciones", "Revisiones", "Casos_Empleado", "Reglas"} <= set(workbook.sheetnames)
    assert set(execution["record_type"]) >= {"CORRIDA", "FUENTE", "REGLA"}
    assert str(tmp_path) not in execution.to_csv(index=False)
