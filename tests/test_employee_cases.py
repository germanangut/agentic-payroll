import pandas as pd
from openpyxl import load_workbook

from agentic_nomina.reporting.cases import build_employee_case_file
from agentic_nomina.reporting.excel import _exception_frame, write_report


def _employee_result(employee_id: str, severity: str = "OK") -> dict[str, object]:
    return {
        "employee_id": employee_id,
        "employee_name_list": f"Employee {employee_id}",
        "employee_name_payroll": f"Employee {employee_id}",
        "employee_status": "E",
        "in_employee_list": True,
        "in_payroll": severity == "OK",
        "outcome": "MATCHED" if severity == "OK" else "MISSING_IN_PAYROLL",
        "severity": severity,
    }


def _social(employee_id: str, severity: str = "OK") -> dict[str, object]:
    return {
        "employee_id": employee_id,
        "employee_name_payroll": f"Employee {employee_id}",
        "health_severity": severity,
        "pension_severity": "OK",
        "days_severity": "OK",
    }


def test_employee_case_file_aggregates_controls_and_worst_severity() -> None:
    employees = {
        "Q1": pd.DataFrame([_employee_result("1"), _employee_result("2", "WARNING")]),
        "Q2": pd.DataFrame([_employee_result("1"), _employee_result("2")]),
    }
    social = pd.DataFrame([_social("1", "REVIEW"), _social("2")])
    loans = pd.DataFrame(
        [
            {
                "employee_id": "1",
                "employee_name_payroll": "Employee 1",
                "period_label": "Q2",
                "severity": "BLOCKING",
            }
        ]
    )

    cases = build_employee_case_file(employees, social, loans=loans).set_index("employee_id")

    assert cases.at["1", "overall_severity"] == "BLOCKING"
    assert cases.at["1", "total_controls"] == 6
    assert cases.at["1", "blocking_controls"] == 1
    assert cases.at["1", "review_controls"] == 1
    assert "EMPLOYEE_LOANS/Q2/balance_movement:BLOCKING" in cases.at["1", "exception_controls"]
    assert cases.at["2", "overall_severity"] == "WARNING"


def test_exception_registry_includes_employee_master_findings() -> None:
    employees = {"Q1": pd.DataFrame([_employee_result("2", "WARNING")])}
    social = pd.DataFrame([_social("2")])

    exceptions = _exception_frame(employees, social, None, None, None)

    assert len(exceptions) == 1
    assert exceptions.iloc[0]["module"] == "EMPLOYEES"
    assert exceptions.iloc[0]["control"] == "MISSING_IN_PAYROLL"


def test_excel_report_contains_employee_cases_and_master_exceptions(tmp_path) -> None:
    output = tmp_path / "reconciliation.xlsx"
    employees = {"Q1": pd.DataFrame([_employee_result("2", "WARNING")])}
    social = pd.DataFrame([_social("2")])

    write_report(output, employees, social)

    workbook = load_workbook(output, read_only=True)
    assert "Casos_Empleado" in workbook.sheetnames
    assert "Excepciones" in workbook.sheetnames
    assert workbook["Casos_Empleado"].max_row == 2
    assert workbook["Excepciones"].max_row == 2
