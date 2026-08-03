from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook

from agentic_nomina.config import load_config
from agentic_nomina.run_manifest import validate_period, validate_run_id
from agentic_nomina.service import run_baseline

DEMO_PERIOD = "2099-01"
DEMO_NOTICE = "SINTÉTICO — DEMOSTRACIÓN — SIN EFECTO FINANCIERO"


def _safe_child(root: Path, *parts: str) -> Path:
    candidate = root.joinpath(*parts)
    if candidate.resolve().parent != root.joinpath(*parts[:-1]).resolve():
        raise ValueError("La ruta de demo propuesta escapa del directorio autorizado.")
    return candidate


def _prepare_output_dir(value: Path) -> Path:
    target = value.absolute()
    if target.exists():
        if target.is_symlink() or not target.is_dir() or any(target.iterdir()):
            raise ValueError("--output-dir debe ser un directorio inexistente o vacío, sin enlaces simbólicos.")
    else:
        target.mkdir(parents=True)
    if target.is_symlink():
        raise ValueError("--output-dir no puede ser un enlace simbólico.")
    return target.resolve()


def _write_payroll(path: Path, days: int, incapacity: int) -> None:
    headers = ["EMPLEADO", "NOMBRE EMPLEADO", "VAR", "SUELDO BASICO", "VAR INC", "INCAPACIDAD"]
    pd.DataFrame([headers, ["999000001", "PERSONA DEMO 001", days, 1_000_000, incapacity, 0]]).to_excel(
        path, header=False, index=False, sheet_name="Hoja1"
    )


def _write_employees(path: Path, include_demo: bool) -> None:
    rows: list[list[object]] = [["NIT", "NOMBRE", "EST", "F. RETIRO"]]
    if include_demo:
        rows.append(["999000001", "PERSONA DEMO 001", "ACTIVO", ""])
    pd.DataFrame(rows).to_excel(path, header=False, index=False, sheet_name="Hoja1")


def _write_pila(path: Path) -> None:
    row: list[object] = [None] * 70
    row[4], row[8] = "999000001", "PERSONA DEMO 001"
    row[35], row[36], row[38] = 60, 2_000_000, 80_000
    row[42], row[43], row[46] = 60, 2_000_000, 80_000
    row[48], row[50], row[51] = 60, 2_000_000, 0
    row[55], row[58], row[61], row[62], row[69] = 60, 2_000_000, 0, 0, 160_000
    pd.DataFrame([row]).to_excel(path, header=False, index=False, sheet_name="Sheet1")


def _write_overtime(path: Path) -> None:
    header = [f"COLUMNA_{index}" for index in range(60)]
    row: list[object] = [0] * 60
    row[0], row[1] = "999000001", "PERSONA DEMO 001"
    pd.DataFrame([[DEMO_NOTICE] + [None] * 59, header, row]).to_excel(
        path, header=False, index=False, sheet_name="TOTAL HORAS EXTRAS"
    )


def _validate_workbook(path: Path, run_id: str, period: str, output_dir: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True)
    required = {"Ejecucion", "Resumen", "Excepciones", "Revisiones", "Casos_Empleado", "Reglas", "Ausencias", "Horas_Extras"}
    if not required <= set(workbook.sheetnames):
        raise ValueError("El libro demo no contiene las hojas esperadas.")
    execution = pd.read_excel(path, sheet_name="Ejecucion")
    if not {run_id} <= set(execution["run_id"].dropna()) or not {period} <= set(execution["period"].dropna()):
        raise ValueError("Ejecucion no conserva run_id y período de la demo.")
    source_rows = execution.loc[execution["record_type"].eq("FUENTE")]
    if source_rows.loc[source_rows["status"].eq("USED"), "sha256"].isna().any():
        raise ValueError("Ejecucion contiene una fuente utilizada sin SHA-256.")
    persisted = execution.to_csv(index=False)
    if str(output_dir) in persisted or ":\\" in persisted:
        raise ValueError("Ejecucion expone una ruta local.")
    return workbook.sheetnames


def run_demo(output_dir: str | Path, *, period: str | None = None, run_id: str | None = None, config_path: str | Path = "config/baseline.yml") -> dict[str, Any]:
    """Generate safe synthetic inputs and invoke the normal reconciliation service."""
    effective_period = validate_period(period or DEMO_PERIOD)
    effective_run_id = validate_run_id(run_id)
    root = _prepare_output_dir(Path(output_dir))
    inputs = _safe_child(root, "inputs")
    results_dir = _safe_child(root, "results")
    inputs.mkdir()
    results_dir.mkdir()
    files = {"payroll_q1": _safe_child(inputs, "payroll-q1.xlsx"), "payroll_q2": _safe_child(inputs, "payroll-q2.xlsx"), "employees_q1": _safe_child(inputs, "employees-q1.xlsx"), "employees_q2": _safe_child(inputs, "employees-q2.xlsx"), "pila": _safe_child(inputs, "pila.xlsx"), "overtime_q1": _safe_child(inputs, "overtime-q1.xlsx"), "overtime_q2": _safe_child(inputs, "overtime-q2.xlsx")}
    _write_payroll(files["payroll_q1"], 28, 2)
    _write_payroll(files["payroll_q2"], 30, 0)
    _write_employees(files["employees_q1"], True)
    _write_employees(files["employees_q2"], False)
    _write_pila(files["pila"])
    _write_overtime(files["overtime_q1"])
    _write_overtime(files["overtime_q2"])
    absence = _safe_child(inputs, "absence-evidence.csv")
    pd.DataFrame([{"employee_id": "999000001", "employee_name": "PERSONA DEMO 001", "absence_type": "INCAPACIDAD", "units": 2, "support_reference": "DEMO-ABS-001"}]).to_csv(absence, index=False)
    manifest_path = _safe_child(root, "manifest.yml")
    manifest_path.write_text(yaml.safe_dump({"schema_version": "1.0", "business_period": effective_period, "run_id": effective_run_id, "sources": {name: str(path.relative_to(root)).replace("\\", "/") for name, path in files.items()}}, sort_keys=False), encoding="utf-8")
    report = _safe_child(results_dir, "agentic-nomina-demo.xlsx")
    result = run_baseline(payroll_q1_path=files["payroll_q1"], payroll_q2_path=files["payroll_q2"], employees_q1_path=files["employees_q1"], employees_q2_path=files["employees_q2"], pila_path=files["pila"], overtime_q1_path=files["overtime_q1"], overtime_q2_path=files["overtime_q2"], absence_evidence_paths=[absence], manifest_path=manifest_path, output_path=report, config=load_config(config_path), business_period=effective_period, run_id=effective_run_id)
    metadata = result["run_metadata"].iloc[0]
    sheets = _validate_workbook(report, str(metadata["run_id"]), str(metadata["period"]), root)
    summary = {"notice": DEMO_NOTICE, "run_id": metadata["run_id"], "business_period": metadata["period"], "preflight_status": metadata["preflight_status"], "manifest": "manifest.yml", "workbook": "results/agentic-nomina-demo.xlsx", "sheets": sheets, "validation": "OK"}
    _safe_child(root, "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary
